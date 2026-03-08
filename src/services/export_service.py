import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text
from data.database import db_manager


class ExportService:
    """
    Service responsible for exporting database tables to Excel.

    Uses SQLAlchemy engine to fetch raw data from all tables and 
    openpyxl to generate a styled spreadsheet.
    """

    def __init__(self, db=None):
        """
        Initializes the export service.

        Args:
            db: Legacy parameter kept for compatibility with main.py.
        """
        self.engine = db_manager.engine

    def export_to_excel(self, filepath: str):
        """
        Exports all relevant application tables to an Excel file.

        Each table is exported to a separate sheet in the workbook.

        Args:
            filepath (str): The destination path for the .xlsx file.
        """
        tables = [
            "interns",
            "venues",
            "documents",
            "observations",
            "meetings",
            "grades",
            "evaluation_criteria",
            "visits"
        ]

        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        if default_sheet:
            wb.remove(default_sheet)

        try:
            with self.engine.connect() as conn:
                for table_name in tables:
                    self._export_table(wb, conn, table_name)

            wb.save(filepath)
            print(f"Exportação concluída: {filepath}")

        except Exception as e:
            print(f"Erro na exportação: {e}")
            raise e

    def _export_table(self, wb, conn, table_name):
        """
        Helper to export a single table to a new sheet in the workbook.
        """
        try:
            # Use raw SQL to fetch all rows from the table
            result = conn.execute(text(f"SELECT * FROM {table_name}"))
            rows = result.fetchall()
            columns = list(result.keys())

            ws = wb.create_sheet(title=table_name.capitalize())

            # Style configuration
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )

            # Add headers
            ws.append(columns)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill

            # Add data rows
            for row in rows:
                ws.append(list(row))

            # Auto-adjust column widths
            for column_cells in ws.columns:
                length = max(len(str(cell.value) or "") for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(
                    length + 2, 50
                )

        except Exception as e:
            # Handle cases where table might not exist or other SQL errors
            print(f"Aviso: Falha ao exportar tabela '{table_name}': {e}")
