import openpyxl
from pathlib import Path
from services.intern_service import InternService
from services.venue_service import VenueService
from services.document_service import DocumentService
from core.models.venue import Venue
from core.models.intern import Intern


class ImportService:
    def __init__(
        self,
        intern_service: InternService,
        venue_service: VenueService,
        document_service: DocumentService,
    ):
        self.intern_service = intern_service
        self.venue_service = venue_service
        self.document_service = document_service
        self._venue_id_map = {}

    def read_file(self, filename: str | Path) -> None:
        path = Path(filename)
        if path.suffix.lower() not in [".xlsx", ".xls"]:
            raise ValueError(
                "O Importador Dedicado aceita apenas arquivos Excel (.xlsx) gerados pelo sistema."
            )

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            sheet_names = wb.sheetnames

            if "Venues" in sheet_names:
                self._process_venues_sheet(wb["Venues"])

            if "Interns" in sheet_names:
                self._process_interns_sheet(wb["Interns"])

        except Exception as e:
            print(f"CRITICAL ERROR NA IMPORTAÇÃO: {e}")
            raise e

    def _sheet_to_dict_list(self, sheet) -> list[dict]:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() for h in rows[0]]
        data = []

        for row in rows[1:]:
            row_data = dict(zip(headers, row))
            clean_data = {k: v for k, v in row_data.items() if v is not None}
            if clean_data:
                data.append(clean_data)

        return data

    def _process_venues_sheet(self, sheet):
        venues_data = self._sheet_to_dict_list(sheet)

        for row in venues_data:
            excel_id = row.get("venue_id")
            name = row.get("venue_name")

            if not name:
                continue

            existing_venue = None
            if excel_id:
                existing_venue = self.venue_service.repo.get_by_id(excel_id)

            if not existing_venue:
                existing_venue = self.venue_service.repo.get_by_name(name)

            venue_obj = Venue(
                venue_id=existing_venue.venue_id if existing_venue else None,
                venue_name=name,
                supervisor_name=row.get("supervisor_name"),
                supervisor_email=row.get("supervisor_email"),
                supervisor_phone=row.get("supervisor_phone"),
            )

            if existing_venue:
                self.venue_service.update_venue(venue_obj)
                real_id = existing_venue.venue_id
            else:
                real_id = self.venue_service.add_new_venue(venue_obj)

            if excel_id and real_id:
                self._venue_id_map[excel_id] = real_id

    def _process_interns_sheet(self, sheet):
        interns_data = self._sheet_to_dict_list(sheet)

        for row in interns_data:
            excel_id = row.get("intern_id")
            name = row.get("name")

            if not name:
                continue

            excel_venue_id = row.get("venue_id")
            real_venue_id = None

            if excel_venue_id:
                if excel_venue_id in self._venue_id_map:
                    real_venue_id = self._venue_id_map[excel_venue_id]
                else:
                    real_venue_id = excel_venue_id

            existing_intern = None
            if excel_id:
                existing_intern = self.intern_service.repo.get_by_id(excel_id)
            if not existing_intern:
                existing_intern = self.intern_service.repo.get_by_name(name)

            intern_obj = Intern(
                intern_id=existing_intern.intern_id if existing_intern else None,
                name=name,
                registration_number=str(row.get("registration_number", "")),
                venue_id=real_venue_id,
                email=row.get("email"),
                start_date=row.get("start_date"),
                end_date=row.get("end_date"),
                working_hours=row.get("working_hours"),
                working_days=row.get("working_days"),
                term=row.get("term", ""),
            )

            if existing_intern:
                self.intern_service.update_intern(intern_obj)
            else:
                self.intern_service.add_new_intern(intern_obj)
